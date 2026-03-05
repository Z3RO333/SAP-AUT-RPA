from __future__ import annotations

from core.common.excel_utils import normalize_column_name


HEADER_ALIASES = {
    "DOCREF": "doc_ref",
    "TIPODOC": "tipo_doc",
    "ORGCOMPRAS": "org_compras",
    "GRUPOCOMPRAS": "grupo_compras",
    "EMPRESA": "empresa",
    "FORNECEDOR": "fornecedor",
    "LIFNR": "fornecedor",
    "MOEDA": "moeda",
    "CONDICAOPAGAMENTO": "condicao_pagamento",
    "ZTERM": "condicao_pagamento",
    "INCOTERMS": "incoterms",
    "TEXTOCABECALHO": "texto_cabecalho",
}

ITEM_ALIASES = {
    "ITEMREF": "item_ref",
    "TEXTOLIVRE": "texto_livre",
    "CATEGORIAITEM": "categoria_item",
    "EPSTP": "categoria_item",
    "CATEGORIACLASSIF": "categoria_classif",
    "KNTTP": "categoria_classif",
    "GRUPOMERCADORIA": "grupo_mercadoria",
    "GRPMERCS": "grupo_mercadoria",
    "WGBEZ": "grupo_mercadoria",
    "PLANTORCENTER": "plant_or_center",
    "CENTRO": "plant_or_center",
    "CEN": "plant_or_center",
    "TAXCODE": "tax_code",
    "MWSKZ": "tax_code",
    "CODIMPOSTO": "tax_code",
    "OBSERVACAOITEM": "observacao_item",
    "CONTARAZAODEFAULT": "conta_razao_default",
    "CONTARAZAO": "conta_razao_default",
    "CONTACONTABIL": "conta_razao_default",
    "SAKTO": "conta_razao_default",
}

SERVICE_ALIASES = {
    "ITEMREF": "item_ref",
    "LINEREF": "line_ref",
    "SERVICELINE": "line_ref",
    "LINHA": "line_ref",
    "NUMEROSERVICO": "numero_servico",
    "NUMSERVICO": "numero_servico",
    "NSERVICO": "numero_servico",
    "SERVICECODE": "numero_servico",
    "SRVPOS": "numero_servico",
    "QUANTIDADE": "quantidade",
    "MENGE": "quantidade",
    "PRECOBRUTO": "preco_bruto",
    "PRECOLIQ": "preco_bruto",
    "VALORTOTAL": "preco_bruto",
    "TBTWR": "preco_bruto",
    "ORDEM": "ordem",
    "AUFNR": "ordem",
    "CENTROCUSTO": "centro_custo",
    "KOSTL": "centro_custo",
    "CONTARAZAO": "conta_razao",
    "CONTACONTABIL": "conta_razao",
    "SAKTO": "conta_razao",
    "DESCRICAOEXTERNA": "descricao_externa",
    "EXTSRVNO": "descricao_externa",
}


def _sheet_to_rows(worksheet, aliases: dict[str, str]) -> list[dict]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [aliases.get(normalize_column_name(value), normalize_column_name(value).lower()) for value in rows[0]]
    output: list[dict] = []
    for row in rows[1:]:
        if not any(value not in (None, "") for value in row):
            continue
        output.append({headers[index]: value for index, value in enumerate(row) if headers[index]})
    return output


def read_me21n_workbook(path: str) -> tuple[list[dict], list[dict], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "O modo de planilha da ME21N requer o pacote 'openpyxl'. Instale-o para ler arquivos XLSX."
        ) from exc

    workbook = load_workbook(path, data_only=True)
    required_sheets = {"CABECALHO", "ITENS", "SERVICOS"}
    missing = sorted(required_sheets - set(workbook.sheetnames))
    if missing:
        raise ValueError(f"O arquivo deve conter as abas obrigatorias: {', '.join(missing)}.")
    header_rows = _sheet_to_rows(workbook["CABECALHO"], HEADER_ALIASES)
    item_rows = _sheet_to_rows(workbook["ITENS"], ITEM_ALIASES)
    service_rows = _sheet_to_rows(workbook["SERVICOS"], SERVICE_ALIASES)
    return header_rows, item_rows, service_rows
